import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Ejecutar git log y capturar commits
resultado = subprocess.run(
    ["git", "log", "--pretty=format:%h|%an|%ad|%s", "--date=short"],
    stdout=subprocess.PIPE,
    stderr=subprocess.PIPE,
    text=True,
    encoding="utf-8",
    errors="replace"
)

if resultado.returncode != 0:
    print("❌ Error ejecutando git log:")
    print(resultado.stderr)
    exit(1)

lineas = resultado.stdout.strip().split("\n")

# Crear libro de Excel
wb = Workbook()
ws = wb.active
ws.title = "Reporte de Avance"

# Título del reporte
ws.merge_cells("A1:E1")
titulo = ws["A1"]
titulo.value = "📊 Reporte de Avance - Git Commits"
titulo.font = Font(size=14, bold=True)
titulo.alignment = Alignment(horizontal="center")

# Encabezados
headers = ["N°", "Hash", "Autor", "Fecha", "Descripción"]
ws.append(headers)

# Estilo de encabezados
for cell in ws[2]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Insertar datos de commits
for i, linea in enumerate(lineas, start=1):
    partes = [p.strip() for p in linea.split("|")]
    if len(partes) == 4:
        hash_, autor, fecha, mensaje = partes
        ws.append([i, hash_, autor, fecha, mensaje])

# Ajustar ancho de columnas
ancho_columnas = [5, 12, 25, 12, 60]
for i, width in enumerate(ancho_columnas, start=1):
    ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width

# Guardar archivo Excel
nombre_archivo = "reporte_avance.xlsx"
wb.save(nombre_archivo)

print(f"✅ Reporte generado correctamente: {nombre_archivo}")
